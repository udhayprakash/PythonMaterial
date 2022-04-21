from __future__ import print_function
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./test.xlsx')

# Get sheet names
print(wb.get_sheet_names())

# Get a sheet by name
sheet = wb.get_sheet_by_name('Sheet1')

# Print the sheet title
sheet.title

# Get currently active sheet
anotherSheet = wb.active

# Check `anotherSheet`
anotherSheet

# Retrieve the value of a certain cell
sheet['A1'].value

# Select element 'B2' of your sheet
c = sheet['B2']

# Retrieve the row number of your element
c.row

# Retrieve the column letter of your element
c.column

# Retrieve the coordinates of the cell
c.coordinate

# Retrieve cell value
sheet.cell(row=1, column=2).value

# Print out values in column 2
for i in range(1, 4):
     print(i, sheet.cell(row=i, column=2).value)
