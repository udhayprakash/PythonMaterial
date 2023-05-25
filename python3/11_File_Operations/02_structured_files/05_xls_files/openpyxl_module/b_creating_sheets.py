#!/usr/bin/python3

import openpyxl


def create_worksheets(path):
    workbook = openpyxl.Workbook()
    print(workbook.sheetnames)  # ['Sheet'] - default

    # Add a new worksheet
    workbook.create_sheet()
    print(workbook.sheetnames)  # ['Sheet', 'Sheet1']

    # Insert a worksheet
    workbook.create_sheet(index=1, title="Second sheet")
    print(workbook.sheetnames)  # ['Sheet', 'Second sheet', 'Sheet1']

    workbook.save(path)


if __name__ == "__main__":
    create_worksheets("b_creating_sheets.xlsx")
