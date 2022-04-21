from openpyxl import load_workbook


def iterating_over_values(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    for value in sheet.iter_rows(
        min_row=1, max_row=3, min_col=1, max_col=3, values_only=True
    ):
        print(value)


if __name__ == "__main__":
    iterating_over_values("f_delete_cols_rows.xlsx")

# ws.max_row will give you the number of rows in a worksheet.

# wb = load_workbook(path, use_iterators=True)
# sheet = wb.worksheets[0]

# row_count = sheet.max_row
# column_count = sheet.max_column
