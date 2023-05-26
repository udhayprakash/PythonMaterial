from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()

ws = wb.active
ws["A1"] = "Line 1\nLine 2\nLine 3"

ws["A1"].alignment = Alignment(wrapText=True)
ws.protection.set_password("test")
wb.save("wrap.xlsx")


def PassProtect(Path, Pass):
    from win32com.client.gencache import EnsureDispatch

    xlApp = EnsureDispatch("Excel.Application")
    xlwb = xlApp.Workbooks.Open(Path)
    xlApp.DisplayAlerts = False
    xlwb.Visible = False
    xlwb.SaveAs(Path, Password=Pass)
    xlwb.Close()
    xlApp.Quit()


PassProtect(
    "wrap.xlsx",
    "DesiredPasswordGoesHere",
)
